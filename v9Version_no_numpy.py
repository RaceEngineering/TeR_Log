import re
import cantools
import csv
from collections import defaultdict
from typing import List, Dict
import matplotlib.pyplot as plt
import pandas as pd
from scipy.io import savemat
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PilImage
import operator

class Signal:
    def __init__(self, dbc_path: str):
        self.db = cantools.database.load_file(dbc_path)
        # Operadores básicos
        self.operations = {
            '+': operator.add,
            '-': operator.sub,
            '*': operator.mul,
            '/': operator.truediv
        }

    def _write_to_csv(self, grouped_decoded: Dict[str, List[float]], csv_final: str):
        with open(csv_final, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            for key, values in grouped_decoded.items():
                writer.writerow([key, values])
        print(f"Decoding completed and saved to {csv_final}")

    def _write_to_xlsx(self, grouped_decoded: Dict[str, List[float]], xlsx_final:str, plot_save_path:str = None):
        df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in grouped_decoded.items()]))
        df.to_excel(xlsx_final, index=False)

        if plot_save_path:
            workbook = load_workbook(xlsx_final)
            sheet = workbook.create_sheet("Plot")

            img = PilImage.open(plot_save_path)
            img = img.convert("RGB")
            img.save("imagen_xlsx.png")

            image = OpenpyxlImage("imagen_xlsx.png")
            sheet.add_image(image, 'A1')
            workbook.save(xlsx_final)
        
        print(f"Decoding and plot saved to {xlsx_final}")
    
    def _write_to_mat(self, grouped_decoded: Dict[str, List[float]], mat_final: str):
        savemat(mat_final, grouped_decoded)
        print(f"Decoding completed and saved to {mat_final}")

    def _write_to_ascii(self, grouped_decoded: Dict[str, List[float]], ascii_final: str):
        with open(ascii_final, mode='w') as ascii_file:
            signals = grouped_decoded.keys()
            ascii_file.write("\t".join(signals) + "\n")
            max_len = max(len(values) for values in grouped_decoded.values())
            for i in range(max_len):
                row = []
                for signal in signals:
                    value = grouped_decoded[signal][i] if i < len(grouped_decoded[signal]) else ""
                    row.append(str(value))
                ascii_file.write("\t".join(row) + "\n")
        print(f"Decoding completed and saved to {ascii_final}")

    def _plot_signals(self, grouped_decoded: Dict[str, List[float]], signal_names: List[str], save_path: str = None):
        plt.figure(figsize=(10, 5))
        for signal_name in signal_names:
            if signal_name in grouped_decoded:
                plt.plot(grouped_decoded[signal_name], label=signal_name)
            else:
                print(f"Signal {signal_name} not found in the decoded data.")
        plt.title("Signals Plot")
        plt.xlabel("Sample Number")
        plt.ylabel("Value")
        plt.legend()
        plt.grid(True)
        
        if save_path:
            plt.savefig(save_path)
            print(f"Plot saved to {save_path}")
        
        plt.show()

    def _apply_operation(self, expression: str, grouped_decoded: Dict[str, List[float]]) -> List[float]:
        """
        Evalúa una expresión matemática que incluye operaciones entre señales decodificadas.
        La expresión debe estar en formato: 'SIGNAL1 + SIGNAL2', 'SIGNAL1 * SIGNAL2', etc.
        """
        tokens = re.split(r'(\s+|\+|\-|\*|\/)', expression)  # Separar la expresión en partes
        stack = []

        for token in tokens:
            token = token.strip()
            if token in self.operations:  # Si es un operador
                stack.append(token)
            elif token in grouped_decoded:  # Si es una señal
                stack.append(grouped_decoded[token])
            else:  # Si es un número
                try:
                    stack.append([float(token)] * len(next(iter(grouped_decoded.values()))))
                except ValueError:
                    pass

        # Evalúa la expresión en la pila
        while len(stack) > 1:
            operand1 = stack.pop(0)
            operator = stack.pop(0)
            operand2 = stack.pop(0)
            result = [self.operations[operator](a, b) for a, b in zip(operand1, operand2)]
            stack.insert(0, result)

        return stack[0]

    def add_operation(self, grouped_decoded: Dict[str, List[float]], expression: str, result_name: str) -> Dict[str, List[float]]:
        """
        Agrega una nueva columna de resultado basada en la expresión dada.
        """
        result = self._apply_operation(expression, grouped_decoded)
        grouped_decoded[result_name] = result  # Añadir el resultado con un nuevo nombre
        return grouped_decoded

    def decode_log(self, log_path: str, output_file: str, output_format: str, signals_to_plot: List[str] = None, plot_save_path: str = None, operations: List[Dict[str, str]] = None):
        """
        Decodifica el log y genera los archivos de salida. Si se proporcionan operaciones, se aplican
        y los resultados se añaden a las señales decodificadas sin modificar los valores originales.
        """
        pattern = r'(can0\s+)(\d+)\s+\[(\d)\]\s+([A-F0-9\s]+)'
        with open(log_path, 'r') as file:
            log = file.read()
        regex = re.compile(pattern)
        grouped_decoded = defaultdict(list)
        
        for match in regex.finditer(log):
            msg = {"data": bytearray.fromhex(match.group(4)), "id": int(match.group(2), 16)}
            log_decode = self.db.decode_message(msg["id"], msg["data"])
            for key, value in log_decode.items():
                if isinstance(value, (int, float)):
                    grouped_decoded[key].append(value)

        # Aplicar operaciones si se proporcionan
        if operations:
            for operation in operations:
                expression = operation['expression']
                result_name = operation.get('result_name', 'Result')
                grouped_decoded = self.add_operation(grouped_decoded, expression, result_name)

        # Guardar en el formato solicitado
        if output_format == 'csv':
            self._write_to_csv(grouped_decoded, output_file)
        elif output_format == 'ascii':
            self._write_to_ascii(grouped_decoded, output_file)
        elif output_format == 'xlsx':
            self._write_to_xlsx(grouped_decoded, output_file, plot_save_path=plot_save_path)
        elif output_format == 'mat':
            self._write_to_mat(grouped_decoded, output_file)
        else:
            print("Unsupported format")
        
        # Graficar si es necesario
        if signals_to_plot:
            self._plot_signals(grouped_decoded, signals_to_plot, save_path=plot_save_path)

# Uso del código
decoder = Signal("./TER.dbc")
decoder.decode_log(
    "RUN0.log", 
    "decoded_log.xlsx", 
    "xlsx", 
    signals_to_plot=["PITCH", "ROLL", "YAW"], 
    plot_save_path="combined_plot.png", 
    operations=[
        {"expression": "PITCH + ROLL", "result_name": "Pitch_Roll_Sum"},
        {"expression": "PITCH - YAW", "result_name": "Pitch_Yaw_Diff"}
    ]
)
