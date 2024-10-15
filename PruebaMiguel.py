import re
import cantools
import csv
from collections import defaultdict
from typing import List, Dict
import matplotlib.pyplot as plt
import pandas as pd
from scipy.io import savemat
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PilImage
import operator
from scipy.interpolate import interp1d

class Signal:
    def __init__(self, dbc_path: str):
        # Cargar el archivo DBC
        self.db = cantools.database.load_file(dbc_path)
        print(f"Loaded DBC: {dbc_path}")

        # Imprimir todos los IDs y sus nombres de mensajes en el DBC
        self._print_message_ids()
        # Operadores básicos
        self.operations = {
            '+': operator.add,
            '-': operator.sub,
            '*': operator.mul,
            '/': operator.truediv
        }
        # Definir la precedencia de operadores
        self.precedence = {
            '+': 2,
            '-': 2,
            '*': 1,
            '/': 1
        }

    def _print_message_ids(self):
        """Imprime los IDs de los mensajes definidos en el DBC."""
        print("Message IDs defined in the DBC:")
        for message in self.db.messages:
            print(f"ID: {message.frame_id} ({message.name})")

    def _write_to_xlsx(self, grouped_decoded: Dict[str, List[float]], timestamps: List[float], xlsx_final: str, plot_save_path: str = None):
        # Crear un DataFrame con los timestamps y las señales
        df = pd.DataFrame({"Timestamp": timestamps})
        for key, values in grouped_decoded.items():
            # Rellenar las señales con NaN en los timestamps donde no tienen valores
            signal_series = pd.Series(values, index=df.index[:len(values)])
            df[key] = signal_series

        # Guardar en Excel
        df.to_excel(xlsx_final, index=False)

        if plot_save_path:
            workbook = load_workbook(xlsx_final)
            sheet = workbook.create_sheet("Plot")

            img = PilImage.open(plot_save_path)
            img = img.convert("RGB")
            img.save("imagen_xlsx.png")

            image = OpenpyxlImage("imagen_xlsx.png")
            sheet.add_image(image, 'A1')
            workbook.save(xlsx_final)
        
        print(f"Decoding and plot saved to {xlsx_final}")

    def _plot_signals(self, grouped_decoded: Dict[str, List[float]], signal_names: List[str], timestamps: List[float], save_path: str = None):
        plt.figure(figsize=(10, 5))
        for signal_name in signal_names:
            if signal_name in grouped_decoded:
                # Usar los timestamps como eje x
                plt.plot(timestamps[:len(grouped_decoded[signal_name])], grouped_decoded[signal_name], label=signal_name)
            else:
                print(f"Signal {signal_name} not found in the decoded data.")
        plt.title("Signals Plot")
        plt.xlabel("Timestamp (s)")
        plt.ylabel("Value")
        plt.legend()
        plt.grid(True)
        
        if save_path:
            plt.savefig(save_path)
            print(f"Plot saved to {save_path}")
        
        plt.show()

    def decode_log(self, log_path: str, output_file: str, output_format: str, signals_to_plot: List[str] = None, plot_save_path: str = None, operations: List[Dict[str, str]] = None):
        # Patrón regex para capturar timestamp, interfaz, ID (3 caracteres) y datos
        pattern = r'\((?P<timestamp>\d+\.\d{6})\)\s+(?P<interface>\w+)\s+(?P<id>[0-9A-F]{3})\s*#\s*(?P<data>[0-9A-F]{2,16})'
        
        # Abrir en modo lectura el log
        with open(log_path, 'r') as file:
            log = file.read()
        
        # Compilar regex
        regex = re.compile(pattern)

        # Diccionario de mensajes decodificados
        grouped_decoded = defaultdict(list)
        timestamps = []

        # Hacer los matches:
        for match in regex.finditer(log):
            msg_id_str = match.group("id")  # ID del mensaje como string
            msg_id = int(msg_id_str, 16)  # Convertir ID a entero

            try:
                # Intentar convertir los datos a bytes
                msg_data = bytearray.fromhex(match.group("data"))
                timestamp = float(match.group("timestamp"))
                timestamps.append(timestamp)
            except ValueError:
                print(f"Error: Los datos '{match.group('data')}' no son válidos como hexadecimal. Se omite este mensaje.")
                pass

            try:
                # Decodificar con el archivo DBC
                log_decode = self.db.decode_message(msg_id, msg_data)
                print(f"Decoded Message ID: {msg_id}, Data: {msg_data.hex()} -> {log_decode}")

                # Agrupar los valores decodificados
                for key, value in log_decode.items():
                    if isinstance(value, (int, float)):
                        grouped_decoded[key].append(value)
                    else:
                        print(f"Warning: Signal '{key}' has non-numeric value '{value}' in message ID {msg_id}. Skipping this value.")
            except KeyError:
                print(f"Warning: Message ID {msg_id_str} (decimal {msg_id}) is not defined in the DBC.")
            except Exception as e:
                print(f"Error decoding message with ID {msg_id_str} (decimal {msg_id}): {e}")

        # Aplicar operaciones si se proporcionan
        if operations:
            for operation in operations:
                expression = operation['expression']
                result_name = operation.get('result_name', 'Result')
                grouped_decoded = self.add_operation(grouped_decoded, expression, result_name)

        # Guardar en el formato solicitado
        if output_format == 'xlsx':
            self._write_to_xlsx(grouped_decoded, timestamps, output_file, plot_save_path=plot_save_path)
        else:
            print("Unsupported format")
        
        # Graficar si es necesario
        if signals_to_plot:
            self._plot_signals(grouped_decoded, signals_to_plot, timestamps, save_path=plot_save_path)


# Uso del código
if __name__ == "__main__":
    decoder = Signal("./TER.dbc")
    decoder.decode_log("RUN4.log", "RUN4.xlsx", "xlsx", signals_to_plot=["rrRPM", "rlRPM", "APPS_AV", "ANGLE"], plot_save_path="combined_plot.png")
