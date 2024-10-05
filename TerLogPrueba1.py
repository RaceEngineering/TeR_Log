import re
import cantools
import csv
from collections import defaultdict
from typing import List, Dict
import matplotlib.pyplot as plt
import pandas as pd
from scipy.io import savemat
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PilImage

class Signal:
    def __init__(self, dbc_path: str):
        self.db = cantools.database.load_file(dbc_path)
    
    def _write_to_csv(self, grouped_decoded: Dict[str, List[float]], csv_final: str):
        with open(csv_final, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["Timestamp"] + list(grouped_decoded.keys()))  # Escribir encabezado
            max_len = max(len(v) for v in grouped_decoded.values())
            for i in range(max_len):
                row = [grouped_decoded['Timestamp'][i]]
                for key in grouped_decoded:
                    if key != 'Timestamp':
                        row.append(grouped_decoded[key][i] if i < len(grouped_decoded[key]) else "")
                writer.writerow(row)
        print(f"Decoding completed and saved to {csv_final}")

    def _write_to_xlsx(self, grouped_decoded: Dict[str, List[float]], xlsx_final: str, plot_save_path: str = None):
        df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in grouped_decoded.items()]))
        df.to_excel(xlsx_final, index=False)

        if plot_save_path:
            workbook = load_workbook(xlsx_final)
            sheet = workbook.create_sheet("Plot")

            img = PilImage.open(plot_save_path)
            img = img.convert("RGB")
            img.save("imagen_xlsx.png")

            image = OpenpyxlImage("imagen_xlsx.png")
            sheet.add_image(image, 'A1')
            workbook.save(xlsx_final)
        
        print(f"Decoding and plot saved to {xlsx_final}")
    
    def _write_to_mat(self, grouped_decoded: Dict[str, List[float]], mat_final: str):
        savemat(mat_final, grouped_decoded)
        print(f"Decoding completed and saved to {mat_final}")

    def _write_to_ascii(self, grouped_decoded: Dict[str, List[float]], ascii_final: str):
        with open(ascii_final, mode='w') as ascii_file:
            signals = grouped_decoded.keys()
            ascii_file.write("\t".join(signals) + "\n")
            max_len = max(len(values) for values in grouped_decoded.values())
            for i in range(max_len):
                row = []
                for signal in signals:
                    value = grouped_decoded[signal][i] if i < len(grouped_decoded[signal]) else ""
                    row.append(str(value))
                ascii_file.write("\t".join(row) + "\n")
        print(f"Decoding completed and saved to {ascii_final}")

    def _plot_signals(self, grouped_decoded: Dict[str, List[float]], signal_names: List[str], save_path: str = None):
        plt.figure(figsize=(10, 5))
        for signal_name in signal_names:
            if signal_name in grouped_decoded:
                plt.plot(grouped_decoded[signal_name], label=signal_name)
            else:
                print(f"Signal {signal_name} not found in the decoded data.")
        plt.title("Signals Plot")
        plt.xlabel("Sample Number")
        plt.ylabel("Value")
        plt.legend()
        plt.grid(True)
        
        if save_path:
            plt.savefig(save_path)
            print(f"Plot saved to {save_path}")
        
        plt.show()

    def decode_log(self, log_path: str, output_file: str, output_format: str, signals_to_plot: List[str] = None, plot_save_path: str = None):
    # Nuevo patrón para incluir el timestamp
        pattern = r'\((\d+\.\d{6})\)\s+(\w+)\s+([0-9A-F]{3})#([0-9A-F]{2,16})'
        with open(log_path, 'r') as file:
            log = file.read()
        regex = re.compile(pattern)
        grouped_decoded = defaultdict(list)
        timestamps = []

        for match in regex.finditer(log):
            timestamp = float(match.group(1))  # Capturar el timestamp
            msg_id = int(match.group(3), 16)  # ID del mensaje en hexadecimal
            msg_data = bytes.fromhex(match.group(4))  # Datos del mensaje

        try:
            log_decode = self.db.decode_message(msg_id, msg_data)
            timestamps.append(timestamp)  # Almacenar el timestamp
            for key, value in log_decode.items():
                if isinstance(value, (int, float)):
                    grouped_decoded[key].append(value)
        except KeyError:
            print(f"Mensaje CAN con ID {msg_id} no encontrado en el archivo DBC.")

        grouped_decoded['Timestamp'] = timestamps  # Añadir los timestamps al diccionario

        # Salida en formato adecuado
        if output_format == 'csv':
            self._write_to_csv(grouped_decoded, output_file)
        elif output_format == 'ascii':
            self._write_to_ascii(grouped_decoded, output_file)
        elif output_format == 'xlsx':
            self._write_to_xlsx(grouped_decoded, output_file, plot_save_path=plot_save_path)
        elif output_format == 'mat':
            self._write_to_mat(grouped_decoded, output_file)
        else:
            print("Unsupported format")
    
        if signals_to_plot:
            self._plot_signals(grouped_decoded, signals_to_plot, save_path=plot_save_path)

# Uso del código adaptado
decoder = Signal("./TER.dbc")
decoder.decode_log("RUN1.log", "decoded_log.xlsx", "xlsx", signals_to_plot=["PITCH", "ROLL", "YAW"], plot_save_path="combined_plot.png")
