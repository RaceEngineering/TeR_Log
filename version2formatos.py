import re #Libreria para python
import cantools #Libreria para decodificar los mensajes del CAN
import csv #Libreria para convertir a csv
from collections import defaultdict
from typing import List, Dict
import matplotlib.pyplot as plt #Libreria para garficar con python
import pandas as pd #Libreria para convertir en xlsx mas facil pasandolo a un Dataframe
from scipy.io import savemat #Libreria para guardar a un .mat
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class Signal:
    def __init__(self, dbc_path: str):
        self.db = cantools.database.load_file(dbc_path)
    
    def _write_to_csv(self, grouped_decoded: Dict[str, List[float]], csv_final: str):
        # Crear el archivo CSV y escribir los datos
        with open(csv_final, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            for key, values in grouped_decoded.items():
                writer.writerow([key, values])
        print(f"Decoding completed and saved to {csv_final}")

    def _write_to_xlsx(self, grouped_decoded: Dict[str, List[float]], xlsx_final:str, plot_save_path:str = None):
         # Convertir el diccionario a un DataFrame de pandas
        df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in grouped_decoded.items()]))
    
        # Escribir el DataFrame a un archivo Excel
        df.to_excel(xlsx_final, index=False)
        if plot_save_path:
            workbook = load_workbook(xlsx_final)
            sheet = workbook.create_sheet("Plot")
            img = Image(plot_save_path)
            sheet.add_image(img, 'A1')  # Inserta la imagen en la celda A1
            workbook.save(xlsx_final)
        print(f"Decoding completed and saved to {xlsx_final}")
    
    def _write_to_mat(self,grouped_decoded: Dict[str, List[float]], mat_final: str):
        savemat(mat_final,grouped_decoded)
        print(f"Decoding completed and saved to {mat_final}")

    
    def _write_to_ascii(self, grouped_decoded: Dict[str, List[float]], ascii_final: str):
    # Crear el archivo ASCII y escribir los datos en un formato de tabla
        with open(ascii_final, mode='w') as ascii_file:
            # Obtener las señales
            signals = grouped_decoded.keys()
             # Escribir el encabezado
            ascii_file.write("\t".join(signals) + "\n")
        
             # Encontrar el número máximo de valores
            max_len = max(len(values) for values in grouped_decoded.values())
            # Escribir los valores de las señales
            for i in range(max_len):
                row = []
                for signal in signals:
                # Si no hay suficiente valor para esa señal, escribir un valor vacío o 0
                    value = grouped_decoded[signal][i] if i < len(grouped_decoded[signal]) else ""
                    row.append(str(value))
                ascii_file.write("\t".join(row) + "\n")
        print(f"Decoding completed and saved to {ascii_final}")


    def _plot_signals(self, grouped_decoded: Dict[str, List[float]], signal_names: List[str], save_path: str = None):
        plt.figure(figsize=(10, 5))

        for signal_name in signal_names:
            if signal_name in grouped_decoded:
                plt.plot(grouped_decoded[signal_name], label=signal_name)
            else:
                print(f"Signal {signal_name} not found in the decoded data.")
        
        plt.title("Signals Plot")
        plt.xlabel("Sample Number")
        plt.ylabel("Value")
        plt.legend()
        plt.grid(True)
        
        if save_path:
            plt.savefig(save_path)  # Guardar la gráfica en un archivo
            print(f"Plot saved to {save_path}")
        
        plt.show()  # Mostrar la gráfica en pantalla

    def decode_log(self, log_path: str, output_file: str, output_format: str, signals_to_plot: List[str] = None, plot_save_path: str = None):
        pattern = r'(can0\s+)(\d+)\s+\[(\d)\]\s+([A-F0-9\s]+)'
       
        # Abrir en modo lectura el log
        with open(log_path, 'r') as file:
            log = file.read()
       
        # Compilar regex
        regex = re.compile(pattern)

        # Dictionario de mensajes decodificados
        grouped_decoded = defaultdict(list)

        # Hacer los matches:
        for match in regex.finditer(log):
            msg = {
                "data": bytearray.fromhex(match.group(4)),
                "id": int(match.group(2), 16)
            }
    
            # Decodificar con el TER.dbc
            log_decode = self.db.decode_message(msg["id"], msg["data"])
    
            # Meter todos los values agrupados en cada key
            for key, value in log_decode.items():
                if isinstance(value, (int, float)):
                    grouped_decoded[key].append(value)
       
        # Das opcion a que eligan formato y llamas a la funcion especifica
        if output_format == 'csv':
            self._write_to_csv(grouped_decoded, output_file)
        elif output_format == 'ascii':
            self._write_to_ascii(grouped_decoded, output_file)
        elif output_format == 'xlsx':
            self._write_to_xlsx(grouped_decoded, output_file)
        elif output_format == 'mat':
            self._write_to_mat(grouped_decoded, output_file)
        else:
            print("Unsupported format")
        
        # Graficar las señales si se especifican
        if signals_to_plot:
            self._plot_signals(grouped_decoded, signals_to_plot, save_path=plot_save_path)

  
# Uso del código
decoder = Signal("./TER.dbc")
decoder.decode_log("RUN0.log", "decoded_log.xlsx", "xlsx", signals_to_plot=["PITCH", "ROLL", "YAW"], plot_save_path="combined_plot.png")
